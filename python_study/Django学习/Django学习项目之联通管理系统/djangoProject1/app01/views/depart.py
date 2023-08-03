from django.shortcuts import render, redirect, HttpResponse

# Create your views here.
from app01 import models

from app01.utils.pagination import Pagination
from app01.utils.bootstrap import BootStrapModelForm
from app01.utils.form import UserModelForm, PrettyModelForm, PrettyEditModelForm
def depart_list(request):
    """部门列表"""
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html()
    }
    return render(request, 'depart_list.html', context)

def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户POST提交过来的数据
    title = request.POST.get("title")

    # 保存到数据库
    queryset = models.Department.objects.create(title=title)

    # 重定向回部门列表
    return redirect("/depart/list/")

def depart_delete(request):
    """部门删除"""
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete() # 删除
    return redirect("/depart/list/")

def depart_edit(request, nid):
    """修改部门"""
    if request.method == "GET":
        # 根据id,获取数据
        row_object = models.Department.objects.filter(id=nid).first()
        print(row_object.id, row_object.title)
        return render(request, 'depart_edit.html', {'row_object': row_object})
    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")

"""
要想读取excel文件需要先 pip install openpyxl
"""
from django.core.files.uploadedfile import InMemoryUploadedFile
from openpyxl import load_workbook
def depart_multi(request):
    """批量上传（excel文件）"""
    # 1. 获取用户上传的文件对象
    file_object = request.FILES.get("excel")
    print(type(file_object))

    # 2. 对象传递给openpyxl，由openpyxl读取文件内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3. 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2): # 从第二行开始
        text = row[0].value
        print(text)
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect('/depart/list/')

